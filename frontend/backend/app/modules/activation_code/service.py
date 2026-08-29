"""Activation code generation. Code becomes the permanent Student ID.
Format: SPK-26-XXXXXX (6 unambiguous alnum chars). DB unique index +
collision-retry guarantee uniqueness under bulk generation."""
import secrets
import uuid

from pymongo.errors import DuplicateKeyError

from app.db.models import ActivationCode, CodeStatus, PromptAudience

# Exclude ambiguous chars (0/O, 1/I/L)
_ALPHABET = "ABCDEFGHJKMNPQRSTUVWXYZ23456789abcdefghjkmnpqrstuvwxyz"
_PREFIX = "SPK-26-"


def _random_code() -> str:
    return _PREFIX + "".join(secrets.choice(_ALPHABET) for _ in range(6))


async def generate_batch(
    count: int,
    created_by: str,
    audience: PromptAudience = PromptAudience.adults,
) -> dict:
    """Generate a batch of codes for one course. Kids and Adults are separate
    courses, so a batch is issued for exactly one of them."""
    if count not in (1, 100, 500, 1000, 5000) and not (1 <= count <= 5000):
        count = max(1, min(count, 5000))
    batch_id = uuid.uuid4().hex[:12]
    created: list[str] = []
    attempts = 0
    max_attempts = count * 5
    while len(created) < count and attempts < max_attempts:
        attempts += 1
        code = _random_code()
        try:
            await ActivationCode(code=code, status=CodeStatus.unused,
                                 batch_id=batch_id, audience=audience).insert()
            created.append(code)
        except DuplicateKeyError:
            continue  # collision: retry with a new code
    return {"batch_id": batch_id, "requested": count, "generated": len(created),
            "audience": audience.value, "codes": created}


async def set_audience(code: str, audience: PromptAudience) -> ActivationCode:
    """Move an unused code to the other course. Refused once it is activated —
    at that point the student's course is fixed and only an admin edit on the
    student record can change it."""
    from app.core.exceptions import ConflictError, NotFoundError

    ac = await ActivationCode.find_one(ActivationCode.code == code.strip())
    if not ac:
        raise NotFoundError("Activation code not found")
    if ac.status == CodeStatus.activated:
        raise ConflictError(
            "This code has already been activated — change the course on the student instead")
    ac.audience = audience
    await ac.save()
    return ac


async def get_valid_unused(code: str) -> ActivationCode:
    from app.core.exceptions import ConflictError, NotFoundError

    ac = await ActivationCode.find_one(ActivationCode.code == code.strip())
    if not ac:
        raise NotFoundError("Activation code not found")
    if ac.status == CodeStatus.blocked:
        raise ConflictError("This activation code is blocked")
    if ac.status == CodeStatus.activated:
        raise ConflictError("This activation code has already been used")
    # "reserved" = allocated to a paid book order awaiting delivery; the buyer
    # activates it once they receive the book, so it is valid for activation.
    return ac


async def mark_activated(ac: ActivationCode, student_id: str) -> None:
    from app.db.base import utcnow

    ac.status = CodeStatus.activated
    ac.activated_at = utcnow()
    ac.activated_student_id = student_id
    await ac.save()


async def set_status(code: str, status: CodeStatus, reason: str | None = None) -> ActivationCode:
    from app.core.exceptions import NotFoundError

    ac = await ActivationCode.find_one(ActivationCode.code == code.strip())
    if not ac:
        raise NotFoundError("Activation code not found")
    ac.status = status
    ac.blocked_reason = reason if status == CodeStatus.blocked else None
    await ac.save()
    return ac


_DELETABLE = {CodeStatus.unused, CodeStatus.blocked}


async def delete_code(code: str, actor: str, reason: str | None = None) -> ActivationCode:
    from app.core.exceptions import ConflictError, NotFoundError

    ac = await ActivationCode.find_one(
        ActivationCode.code == code.strip(), ActivationCode.is_archived == False  # noqa: E712
    )
    if not ac:
        raise NotFoundError("Activation code not found")
    if ac.status not in _DELETABLE:
        raise ConflictError("Only unused or blocked codes can be deleted")
    ac.archive(actor, reason or "deleted by admin")
    await ac.save()
    return ac


async def bulk_delete(codes: list[str], actor: str, reason: str | None = None) -> dict:
    from app.core.exceptions import AppError

    deleted: list[str] = []
    skipped: list[dict] = []
    for code in codes:
        try:
            await delete_code(code, actor, reason)
            deleted.append(code.strip())
        except AppError as exc:
            skipped.append({"code": code.strip(), "reason": exc.message})
    return {"deleted": len(deleted), "codes": deleted, "skipped": skipped}


_MAX_IMPORT = 5000


def _parse_import_xlsx(raw: bytes) -> list[tuple[str, PromptAudience | None, str | None]]:
    """Rows from the Export Excel layout: code + course columns. Other columns
    are ignored — we never restore activated/student state from a spreadsheet."""
    from io import BytesIO

    from openpyxl import load_workbook

    from app.core.exceptions import ValidationAppError

    try:
        wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationAppError("Could not read the Excel file") from exc
    ws = wb.active
    if ws is None:
        raise ValidationAppError("The Excel file has no sheet")
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise ValidationAppError("The Excel file is empty")
    names = [str(c).strip().lower() if c is not None else "" for c in header]
    if "code" not in names:
        raise ValidationAppError("Excel must have a 'code' column (same as Export Excel)")
    code_i = names.index("code")
    course_i = names.index("course") if "course" in names else (
        names.index("audience") if "audience" in names else None
    )
    parsed: list[tuple[str, PromptAudience | None, str | None]] = []
    for row in rows_iter:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        code = str(row[code_i]).strip() if code_i < len(row) and row[code_i] is not None else ""
        if not code or code.lower() == "none":
            parsed.append(("", None, "missing code"))
            continue
        audience: PromptAudience | None = PromptAudience.adults
        if course_i is not None and course_i < len(row) and row[course_i] not in (None, ""):
            course = str(row[course_i]).strip().lower()
            try:
                audience = PromptAudience(course)
            except ValueError:
                parsed.append((code, None, f"unknown course '{row[course_i]}'"))
                continue
        parsed.append((code, audience, None))
    return parsed


async def import_from_xlsx(raw: bytes) -> dict:
    """Insert unused codes from an Export Excel file. Existing codes are left
    untouched and returned in `existing` so admin can see the flags."""
    from app.core.exceptions import ValidationAppError

    parsed = _parse_import_xlsx(raw)
    if not parsed:
        raise ValidationAppError("No data rows in the spreadsheet")
    if len(parsed) > _MAX_IMPORT:
        raise ValidationAppError(f"At most {_MAX_IMPORT} rows per import")

    batch_id = uuid.uuid4().hex[:12]
    imported: list[str] = []
    existing: list[str] = []
    invalid: list[dict] = []
    seen: set[str] = set()
    for code, audience, err in parsed:
        if err:
            invalid.append({"code": code, "reason": err})
            continue
        if code in seen:
            existing.append(code)
            continue
        seen.add(code)
        if await ActivationCode.find_one(ActivationCode.code == code):
            existing.append(code)
            continue
        try:
            await ActivationCode(
                code=code, status=CodeStatus.unused,
                batch_id=batch_id, audience=audience or PromptAudience.adults,
            ).insert()
            imported.append(code)
        except DuplicateKeyError:
            existing.append(code)
    return {
        "batch_id": batch_id,
        "imported": len(imported),
        "codes": imported,
        "existing": existing,
        "invalid": invalid,
    }
